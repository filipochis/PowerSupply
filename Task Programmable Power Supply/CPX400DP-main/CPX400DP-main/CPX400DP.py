import logging
import time
from datetime import timedelta
import xlrd
import os
import serial
import serial.tools.list_ports
import sys
import openpyxl
import tkinter as tk
import tkinter.ttk as ttk
from tkinter import filedialog
from tkinter import ttk
import typing
logger = logging.getLogger()
from ttkthemes import ThemedStyle


__version__ = '1.1.10'

class CPX400DP:
    documentation_url = 'https://resources.aimtti.com/manuals/CPX400DP_Instruction_Manual-Iss1.pdf#page=28'
    
    def __init__(self) -> None:
        self.com_port = None
        self.serialConnection = None
    
    def __str__(self):
        return str(vars(self))
    
    def connect(self, com_port: str, baudrate: int = 9600) -> None:
        """
        Connect to a serial com_port with the specified baud rate.
        
        Parameters:
        - com_port (str): The name or identifier of the serial com_port.
        - baudrate (int, optional): The baud rate for the serial connection. Default is 9600.
        
        Raises:
        - Exception: If an error occurs while connecting to the specified com_port.
        """
        try:
            self.serialConnection = serial.Serial(com_port, baudrate)
            self.com_port = com_port
            logger.debug(f'CPX400DP: Connected to "{com_port}" with speed "{baudrate}".')
        except Exception as e:
            logger.warning(f'CPX400DP: An error occurred while connecting to com_port due to {repr(e)}.')
            raise e
    
    def autoConnect(self) -> None:
        """
        Automatically connect to a CPX400DP device if 1 is found.
        
        Raises:
        - ValueError: If no CPX400DP supplies are detected or if multiple supplies are detected without specifying a specific com com_port.
        - Exception: If an error occurs during the automatic connection process.
        """
        try:
            comports = list(serial.tools.list_ports.comports())
            power_supply_comports = [comport for comport in comports if comport.hwid[:21] == 'USB VID:PID=103E:0460']
            
            power_supply_connected_count = len(power_supply_comports)
            match power_supply_connected_count:
                case 0:
                    raise ConnectionError('No CPX400DP supplies detected.')
                case 1:
                    power_supply_comport = power_supply_comports[0]
                    com_port = power_supply_comport.device
                    self.connect(com_port)
                    self.com_port = com_port
                case _:
                    raise ConnectionError('Multiple CPX400DP supplies detected. One must be specified')
        except ValueError as e:
            logger.error(str(e))
            raise e
        except Exception as e:
            if self.com_port is None:
                logger.error(f'CPX400DP: An error occurred while automatically connecting to device due to {repr(e)}')
            else:
                logger.error(f'CPX400DP {self.com_port}: An error occurred while automatically connecting to device due to {repr(e)}')
            raise e
    
    def disconnect(self) -> None:
        """
        Disconnect from the currently connected serial com_port.
        
        Raises:
        - Exception: If an error occurs while terminating the connection.
        """
        try:
            self.serialConnection.close()
            logger.debug(f'CPX400DP {self.com_port}: Terminated connection.')
        except Exception as e:
            logger.warning(f'CPX400DP {self.com_port}: An error occured while terminating connection due to {repr(e)}.')
            raise e
    
    def setVoltage(self, voltage: float, channel: int = 1) -> None:
        """
        Set the voltage for a specific channel.
        
        Parameters:
        - voltage (float): The voltage value to set for the channel.
        - channel (int, optional): The channel number. Default is 1.
        
        Raises:
        - Exception: If an error occurs while setting the voltage.
        """
        try:
            self.serialConnection.write(bytes(f'V{channel} {voltage}\n'.encode('utf-8')))
            logger.debug(f'CPX400DP {self.com_port}: Set voltage for channel {channel} to {voltage}V.')
        except Exception as e:
            logger.warning(f'CPX400DP {self.com_port}: An error occured while setting voltage due to {repr(e)}')
            raise e
    
    def setCurrent(self, current: float, channel: int = 1) -> None:
        """
        Set the current for a specific channel.
        
        Parameters:
        - current (float): The current value to set for the channel.
        - channel (int, optional): The channel number. Default is 1.
        
        Raises:
        - Exception: If an error occurs while setting the current.
        """
        try:
            self.serialConnection.write(bytes(f'I{channel} {current}\n'.encode('utf-8')))
            logger.debug(f'CPX400DP {self.com_port}: Set current for channel {channel} to {current}A.')
        except Exception as e:
            logger.warning(f'CPX400DP {self.com_port}: An error occured while setting current due to {repr(e)}')
            raise e
    
    def enableOutput(self, channel: int = 1) -> None:
        """
        Enable the output for a specific channel.
        
        Parameters:
        - channel (int, optional): The channel number. Default is 1.
        
        Raises:
        - Exception: If an error occurs while enabling the channel.
        """
        try:
            self.serialConnection.write(bytes(f'OP{channel} 1\n'.encode('utf-8')))
            logger.debug(f'CPX400DP {self.com_port}: Enabled channel {channel}.')
        except Exception as e:
            logger.warning(f'CPX400DP {self.com_port}: An error occured while enabling channel {channel} due to {repr(e)}')
            raise e
    
    def disableOutput(self, channel: int = 1) -> None:
        """
        Disables the output for a specific channel.
        
        Parameters:
        - channel (int, optional): The channel number. Default is 1.
        
        Raises:
        - Exception: If an error occurs while enabling the channel.
        """
        try:
            self.serialConnection.write(bytes(f'OP{channel} 0\n'.encode('utf-8')))
            logger.debug(f'CPX400DP {self.com_port}: Disabled channel {channel}.')
        except Exception as e:
            logger.warning(f'CPX400DP {self.com_port}: An error occured while disabling channel {channel} due to {repr(e)}')
            raise e
    
    def getSetVoltage(self, channel: int = 1) -> float:
        """
        Get the set voltage for a specific channel.
        
        Parameters:
        - channel (int, optional): The channel number. Default is 1.
        
        Returns:
        float: The set voltage for the specified channel.
        
        Raises:
        - Exception: If an error occurs while getting the set voltage.
        """
        try:
            self.serialConnection.write(bytes(f'V{channel}?\n'.encode('utf-8')))
            logger.debug(f'CPX400DP {self.com_port}: Got set voltage for channel {channel}.')
            return float(self.serialConnection.readline()[3:].decode('utf-8').strip())
        except Exception as e:
            logger.warning(f'CPX400DP {self.com_port}: An error occured while getting set voltage due to {repr(e)}')
            raise e
    
    def getSetCurrent(self, channel: int = 1) -> float:
        """
        Get the set current for a specific channel.
        
        Parameters:
        - channel (int, optional): The channel number. Default is 1.
        
        Returns:
        float: The set current for the specified channel.
        
        Raises:
        - Exception: If an error occurs while getting the set current.
        """
        try:
            self.serialConnection.write(bytes(f'I{channel}?\n'.encode('utf-8')))
            logger.debug(f'CPX400DP {self.com_port}: Got set current for channel {channel}.')
            return float(self.serialConnection.readline()[3:].decode('utf-8').strip())
        except Exception as e:
            logger.warning(f'CPX400DP {self.com_port}: An error occured while getting set current due to {repr(e)}')
            raise e
    
    def getOutputVoltage(self, channel: int = 1) -> float:
        """
        Get the output voltage for a specific channel.
        
        Parameters:
        - channel (int, optional): The channel number. Default is 1.
        
        Returns:
        float: The output voltage for the specified channel.
        
        Raises:
        - Exception: If an error occurs while getting the output voltage.
        """
        try:
            self.serialConnection.write(bytes(f'V{channel}O?\n'.encode('utf-8')))
            voltage = float(''.join(filter(lambda char: char.isdigit() or char == '.', self.serialConnection.readline().decode('utf-8'))))
            logger.debug(f'CPX400DP {self.com_port}: Got output voltage for channel {channel}.')
            return voltage
        except Exception as e:
            logger.warning(f'CPX400DP {self.com_port}: An error occured while getting output voltage due to {repr(e)}')
            raise e
    
    def getOutputCurrent(self, channel: int = 1) -> float:
        """
        Get the output current for a specific channel.
        
        Parameters:
        - channel (int, optional): The channel number. Default is 1.
        
        Returns:
        float: The output voltage for the specified channel.
        
        Raises:
        - Exception: If an error occurs while getting the output voltage.
        """
        try:
            self.serialConnection.write(bytes(f'I{channel}O?\n'.encode('utf-8')))
            current = float(''.join(filter(lambda char: char.isdigit() or char == '.', self.serialConnection.readline().decode('utf-8'))))
            logger.debug(f'CPX400DP {self.com_port}: Got output current for channel {channel}.')
            return current
        except Exception as e:
            logger.warning(f'CPX400DP {self.com_port}: An error occured while getting output current due to {repr(e)}')
            raise e
    
    def getOutputStatus(self, channel: int = 1) -> typing.Literal["ON", "OFF"]:
        """
        Get the output status for a specific channel.
        
        This function sends a command to query and retrieve the output status for the specified channel.
        
        Parameters:
        - channel (int, optional): The channel number. Default is 1.
        
        Returns:
        Literal["ON", "OFF"]: The output status ("ON" or "OFF") for the specified channel.
        
        Raises:
        - Exception: If an error occurs while getting the output status.
        """
        try:
            self.serialConnection.write(bytes(f'OP{channel}?\n'.encode('utf-8')))
            status = "ON" if int(self.serialConnection.readline().decode('utf-8')[0]) == 1 else "OFF"
            logger.debug(f'CPX400DP {self.com_port}: Got output status for channel {channel}.')
            return status
        except Exception as e:
            logger.warning(f'CPX400DP {self.com_port}: An error occured while getting output current due to {repr(e)}')
            raise e
    
    def lock(self) -> None:
        """
        Lock the settings.
        
        Raises:
        - Exception: If an error occurs while locking the settings.
        """
        try:
            self.serialConnection.write(bytes('IFLOCK\n'.encode('utf-8')))
            logger.debug(f'CPX400DP {self.com_port}: Locked settings.')
        except Exception as e:
            logger.warning(f'CPX400DP {self.com_port}: An error occured while locking settings due to {repr(e)}')
    
    def unlock(self) -> None:
        """
        Unlock the settings.
        
        Raises:
        - Exception: If an error occurs while locking the settings.
        """
        try:
            self.serialConnection.write(bytes('IFUNLOCK\n'.encode('utf-8')))
            logger.debug(f'CPX400DP {self.com_port}: Unlocked settings.')
        except Exception as e:
            logger.warning(f'CPX400DP {self.com_port}: An error occured while unlocking settings due to {repr(e)}')
    
    def getIdentification(self) -> typing.Dict[str, str]:
        """
        Get identification information.
        
        Returns:
        Dict[str, str]: A dictionary containing identification information with keys 'name', 'model', 'sn', and 'version'.
        
        Raises:
        - Exception: If an error occurs while getting the identification information.
        """
        try:
            self.serialConnection.write(bytes(f'*IDN?\n'.encode('utf-8')))
            data = self.serialConnection.readline().decode('utf-8').split(', ')
            data_dict = {'name': data[0], 'model': data[1], 'sn': data[2], 'version': data[3][:-2]}
            logger.debug(f'CPX400DP {self.com_port}: Got identification info.')
            return data_dict
        except Exception as e:
            logger.warning(f'CPX400DP {self.com_port}: An error occured while getting output current due to {repr(e)}')
            raise e
   

class ONOFF:
   
    def __init__(self):
        # Initialize the connection to the power supply
        logger.info("Power supply initialized.")

    def enableOutput(self, channel):
        logger.info(f"Channel {channel} turned on.")

    def disableOutput(self, channel):
        logger.info(f"Channel {channel} turned off.")

    def autoConnect(self):
        # Auto-connect to the power supply
        logger.info("Auto-connecting to the power supply.")

def main():
    CPX = CPX400DP()
    CPX.autoConnect()
    excel_file = r"C:\Users\uif55698\OneDrive - Continental AG\Desktop\qw .xlsx"  # Adjust the path to your Excel file

    while True:
        try:
            logger.debug("Attempting to load Excel file.")
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            channel1_state = sheet['E2'].value
            channel2_state = sheet['F2'].value
            duration = sheet['G2'].value  # Duration in seconds
            logger.debug(f"Channel 1 state from Excel: {channel1_state}, Channel 2 state from Excel: {channel2_state}")
            
            if channel1_state == "On":
                CPX.enableOutput(channel=1)
            elif channel1_state == "Off":
                CPX.disableOutput(channel=1)
                
            if channel2_state == "On":
                CPX.enableOutput(channel=2)
            elif channel2_state == "Off":
                CPX.disableOutput(channel=2)
                
            logger.info("Power supply states updated from Excel.")

            if isinstance(duration, (int, float)):
                logger.info(f"Waiting for {duration} seconds.")
                time.sleep(duration)
                CPX.disableOutput(channel=1)
                CPX.disableOutput(channel=2)
                logger.info("Channels turned off after the specified duration.")
            else:
                logger.error(f"Invalid duration value in Excel: {duration}. Please enter a valid number.")

        except Exception as e:
            logger.error(f"Error updating power supply states: {e}")

        time.sleep(5)  # Check every 5 seconds

if __name__ == '__main__':
    main()

if __name__ == '__main__':
    # Clear latest.log if it exists
    if os.path.exists('latest.log'):
        open('latest.log', 'w').close()

    # File handler
    file_handler = logging.FileHandler('latest.log', encoding='utf-8')
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(logging.Formatter('%(asctime)s.%(msecs)03d %(levelname)s: %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
    logger.addHandler(file_handler)

    # Console handler
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(logging.Formatter('%(asctime)s.%(msecs)03d %(levelname)s: %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
    logger.addHandler(console_handler)

    # Set the overall logging level
    logger.setLevel(logging.DEBUG)

    try:
        main()
    except Exception as e:
        logger.fatal(f'The script could no longer continue due to {repr(e)}.')
        input('Press any key to exit.')
        exit(1)

class GuiAppWithExcelChannel1:
    def __init__(self, power_supply_instance, excel_path, master=None):
        self.power_supply_instance = power_supply_instance
        self.excel_path = excel_path
        
        # Build UI
        self.window = tk.Tk() if master is None else tk.Toplevel(master)
        self.window.configure(height=800, width=800)
        self.window.resizable(True, True)
        self.window.title("Power Supply Control")
        
         # Title
        self.title_label = tk.Label(self.window, text="Channel 1 Configuration - Temporary Settings", font=("Helvetica", 12, "bold"), bg="#ff6f00")
        self.title_label.pack(pady=20)
        
        # Apply Button
        self.apply_button = ttk.Button(self.window, text="Apply Settings", command=self.apply_temporary_settings)
        self.apply_button.pack(pady=60)
    
    def apply_temporary_settings(self):
        
        # Read data from Excel file
        try:
            workbook = openpyxl.load_workbook(self.excel_path)
            sheet = workbook.active
            channel1_voltage = float(sheet.cell(row=2, column=1).value) if sheet.cell(row=1, column=1).value is not None else 0.0
            channel1_current = float(sheet.cell(row=2, column=2).value) if sheet.cell(row=1, column=2).value is not None else 0.0
        except Exception as e:
            logger.error(f"Error reading data from Excel file: {repr(e)}")
            return
        
        
        # Apply temporary settings to channel 1 of the power supply
        self.power_supply_instance.setVoltage(channel1_voltage, channel=1)
        self.power_supply_instance.setCurrent(channel1_current, channel=1)
        
        # Schedule a function call to revert to initial values after 10 seconds
        self.window.after(10000, self.revert_to_initial_values)
        
        logger.info("Temporary settings applied.")
    
    def revert_to_initial_values(self):
        # Revert to initial values for channel 1 of the power supply
        # You need to define the initial values and how to set them here
        initial_voltage = 0.0
        initial_current = 0.0
        self.power_supply_instance.setVoltage(initial_voltage, channel=1)
        self.power_supply_instance.setCurrent(initial_current, channel=1)
        
        logger.info("Reverted to initial values.")

def main():
    # Specify Excel file path
    excel_path = r"C:\Users\uif55698\OneDrive - Continental AG\Desktop\qw .xlsx"
    
    CPX = CPX400DP()
    CPX.autoConnect()
    GUI = GuiAppWithExcelChannel1(CPX, excel_path)
    GUI.window.mainloop()

if __name__ == '__main__':
    # Clear latest.log if it exists
    if os.path.exists('latest.log'):
        open('latest.log', 'w').close()
    
    # File handler
    file_handler = logging.FileHandler('latest.log', encoding='utf-8')
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(logging.Formatter('%(asctime)s.%(msecs)03d %(levelname)s: %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
    logger.addHandler(file_handler)
    
    # Console handler
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(logging.Formatter('%(asctime)s.%(msecs)03d %(levelname)s: %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
    logger.addHandler(console_handler)
    
    # Set the overall logging level
    logger.setLevel(logging.DEBUG)
    
    try:
        main()
    except Exception as e:
        logger.fatal(f'The script could no longer continue due to {repr(e)}.')
        input('Press any key to exit.')
        exit(1)

        # a doua clasa: ia timpu din excel si il pune pe power supply pt 10 secunde (pe canalul 2)

class GuiAppWithExcelChannel2:
    def __init__(self, power_supply_instance, excel_path, master=None):
        self.power_supply_instance = power_supply_instance
        self.excel_path = excel_path
        

        # Build UI
        self.window = tk.Tk() if master is None else tk.Toplevel(master)
        self.window.configure(height=800, width=800)
        self.window.resizable(True, True)
        self.window.title("Power Supply Control")
        
         # Title
        self.title_label = tk.Label(self.window, text="Channel 2 Configuration - Temporary Settings", font=("Helvetica", 12, "bold"), bg="#3498db")
        self.title_label.pack(pady=20)
        
        # Apply Button
        self.apply_button = ttk.Button(self.window, text="Apply Settings", command=self.apply_temporary_settings)
        self.apply_button.pack(pady=60)
    
    
    def apply_temporary_settings(self):
        # Read data from Excel file
        try:
            workbook = openpyxl.load_workbook(self.excel_path)
            sheet = workbook.active
            channel2_voltage = float(sheet.cell(row=2, column=3).value) if sheet.cell(row=2, column=3).value is not None else 0.0
            channel2_current = float(sheet.cell(row=2, column=4).value) if sheet.cell(row=2, column=4).value is not None else 0.0
        except Exception as e:
            logger.error(f"Error reading data from Excel file: {repr(e)}")
            return
        
        # Apply temporary settings to channel 2 of the power supply
        self.power_supply_instance.setVoltage(channel2_voltage, channel=2)
        self.power_supply_instance.setCurrent(channel2_current, channel=2)
        
        # Schedule a function call to revert to initial values after 10 seconds
        self.window.after(10000, self.revert_to_initial_values)
        
        logger.info("Temporary settings applied.")
    
    def revert_to_initial_values(self):
        # Revert to initial values for channel 2 of the power supply
        # You need to define the initial values and how to set them here
        initial_voltage = 0.0
        initial_current = 0.0
        self.power_supply_instance.setVoltage(initial_voltage, channel=2)
        self.power_supply_instance.setCurrent(initial_current, channel=2)
        
        logger.info("Reverted to initial values.")

def main():
    # Specify Excel file path
    excel_path = r"C:\Users\uif55698\OneDrive - Continental AG\Desktop\qw .xlsx"
    
    CPX = CPX400DP()
    CPX.autoConnect()
    GUI = GuiAppWithExcelChannel2(CPX, excel_path)
    GUI.window.mainloop()

if __name__ == '__main__':
    # Clear latest.log if it exists
    if os.path.exists('latest.log'):
        open('latest.log', 'w').close()
    
    # File handler
    file_handler = logging.FileHandler('latest.log', encoding='utf-8')
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(logging.Formatter('%(asctime)s.%(msecs)03d %(levelname)s: %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
    logger.addHandler(file_handler)
    
    # Console handler
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(logging.Formatter('%(asctime)s.%(msecs)03d %(levelname)s: %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
    logger.addHandler(console_handler)
    
    # Set the overall logging level
    logger.setLevel(logging.DEBUG)
    
    try:
        main()
    except Exception as e:
        logger.fatal(f'The script could no longer continue due to {repr(e)}.')
        input('Press any key to exit.')
        exit(1)

#ia timpu din excel si il pune pe canalul 1

class GuiAppWithExcelChannel1:
    def __init__(self, power_supply_instance, excel_path, master=None):
        self.power_supply_instance = power_supply_instance
        self.excel_path = excel_path
        
       # Build UI
        self.window = tk.Tk() if master is None else tk.Toplevel(master)
        self.window.configure(height=800, width=800)
        self.window.resizable(True, True)
        self.window.title("Power Supply Control")
       
               # Title
        self.title_label = tk.Label(self.window, text="Channel 1 - Value without reverting", font=("Helvetica", 12, "bold"), bg="#f1c40f")
        self.title_label.pack(pady=20)
        
        # Apply Button
        self.apply_button = ttk.Button(self.window, text="Apply Settings", command=self.apply_settings)
        self.apply_button.pack(pady=20)
    
    def apply_settings(self):
        # Read data from Excel file
        try:
            workbook = openpyxl.load_workbook(self.excel_path)
            sheet = workbook.active
            channel1_voltage = float(sheet.cell(row=2, column=1).value) if sheet.cell(row=2, column=1).value is not None else 0.0
            channel1_current = float(sheet.cell(row=2, column=2).value) if sheet.cell(row=2, column=2).value is not None else 0.0
        except Exception as e:
            logger.error(f"Error reading data from Excel file: {repr(e)}")
            return
        
        # Apply settings to channel 1 of the power supply
        self.power_supply_instance.setVoltage(channel1_voltage, channel=1)
        self.power_supply_instance.setCurrent(channel1_current, channel=1)
        
        logger.info("Settings applied.")

def main():
    # Specify Excel file path
    excel_path = r"C:\Users\uif55698\OneDrive - Continental AG\Desktop\qw .xlsx"
    
    CPX = CPX400DP()
    CPX.autoConnect()
    GUI = GuiAppWithExcelChannel1(CPX, excel_path)
    GUI.window.mainloop()

if __name__ == '__main__':
    # Clear latest.log if it exists
    if os.path.exists('latest.log'):
        open('latest.log', 'w').close()
    
    # File handler
    file_handler = logging.FileHandler('latest.log', encoding='utf-8')
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(logging.Formatter('%(asctime)s.%(msecs)03d %(levelname)s: %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
    logger.addHandler(file_handler)
    
    # Console handler
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(logging.Formatter('%(asctime)s.%(msecs)03d %(levelname)s: %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
    logger.addHandler(console_handler)
    
    # Set the overall logging level
    logger.setLevel(logging.DEBUG)
    
    try:
        main()
    except Exception as e:
        logger.fatal(f'The script could no longer continue due to {repr(e)}.')
        input('Press any key to exit.')
        exit(1)

#ia timpu din excel si il pune pe canalul 2

class GuiAppWithExcelChannel1:
    def __init__(self, power_supply_instance, excel_path, master=None):
        self.power_supply_instance = power_supply_instance
        self.excel_path = excel_path
        
       # Build UI
        self.window = tk.Tk() if master is None else tk.Toplevel(master)
        self.window.configure(height=800, width=800)
        self.window.resizable(True, True)
        self.window.title("Power Supply Control")
       
               # Title
        self.title_label = tk.Label(self.window, text="Channel 2 - Value without reverting", font=("Helvetica", 12, "bold"), bg="#2ecc71")
        self.title_label.pack(pady=20)
        
        # Apply Button
        self.apply_button = ttk.Button(self.window, text="Apply Settings", command=self.apply_settings)
        self.apply_button.pack(pady=20)
    
    def apply_settings(self):
        
        # Read data from Excel file
        try:
            workbook = openpyxl.load_workbook(self.excel_path)
            sheet = workbook.active
            channel1_voltage = float(sheet.cell(row=2, column=3).value) if sheet.cell(row=2, column=3).value is not None else 0.0
            channel1_current = float(sheet.cell(row=2, column=4).value) if sheet.cell(row=2, column=4).value is not None else 0.0
        except Exception as e:
            logger.error(f"Error reading data from Excel file: {repr(e)}")
            return
        
        # Apply settings to channel 2 of the power supply
        self.power_supply_instance.setVoltage(channel1_voltage, channel=2)
        self.power_supply_instance.setCurrent(channel1_current, channel=2)
        
        logger.info("Settings applied.")

def main():
    # Specify Excel file path
    excel_path = r"C:\Users\uif55698\OneDrive - Continental AG\Desktop\qw .xlsx"
    
    CPX = CPX400DP()
    CPX.autoConnect()
    GUI = GuiAppWithExcelChannel1(CPX, excel_path)
    GUI.window.mainloop()

if __name__ == '__main__':
    # Clear latest.log if it exists
    if os.path.exists('latest.log'):
        open('latest.log', 'w').close()
    
    # File handler
    file_handler = logging.FileHandler('latest.log', encoding='utf-8')
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(logging.Formatter('%(asctime)s.%(msecs)03d %(levelname)s: %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
    logger.addHandler(file_handler)
    
    # Console handler
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(logging.Formatter('%(asctime)s.%(msecs)03d %(levelname)s: %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
    logger.addHandler(console_handler)
    
    # Set the overall logging level
    logger.setLevel(logging.DEBUG)
    
    try:
        main()
    except Exception as e:
        logger.fatal(f'The script could no longer continue due to {repr(e)}.')
        input('Press any key to exit.')
        exit(1)

  # ON/OFF buttons pentru cele 2 canale    
class GuiAppWithButtons:
    def __init__(self, power_supply_instance, master=None):
        self.power_supply_instance = power_supply_instance
        
        # Build UI
        self.window = tk.Tk() if master is None else tk.Toplevel(master)
        self.window.configure(height=200, width=500)
        self.window.resizable(False, False)
        self.window.title("Power Supply Control")
        
       # Title
        self.title_label = tk.Label(self.window, text="ON/OFF Buttons", font=("Arial", 20, "bold"), fg="#2ecc71")
        self.title_label.place(relx=0.5, rely=0.1, anchor=tk.CENTER)
        
        # Channel 1 On Button
        self.channel1_on_button = ttk.Button(self.window, text="Channel 1 On", command=self.channel1_on)
        self.channel1_on_button.place(x=90, y=100)
        
        # Channel 1 Off Button
        self.channel1_off_button = ttk.Button(self.window, text="Channel 1 Off", command=self.channel1_off)
        self.channel1_off_button.place(x=180, y=100)
        
        # Channel 2 On Button
        self.channel2_on_button = ttk.Button(self.window, text="Channel 2 On", command=self.channel2_on)
        self.channel2_on_button.place(x=290, y=100)
        
        # Channel 2 Off Button
        self.channel2_off_button = ttk.Button(self.window, text="Channel 2 Off", command=self.channel2_off)
        self.channel2_off_button.place(x=380, y=100)
        
        # Other UI elements (Entry fields, labels, etc.) can be added here...
    
    def channel1_on(self):
        self.power_supply_instance.enableOutput(channel=1)
        logger.info("Channel 1 turned on.")
    
    def channel1_off(self):
        self.power_supply_instance.disableOutput(channel=1)
        logger.info("Channel 1 turned off.")
    
    def channel2_on(self):
        self.power_supply_instance.enableOutput(channel=2)
        logger.info("Channel 2 turned on.")
    
    def channel2_off(self):
        self.power_supply_instance.disableOutput(channel=2)
        logger.info("Channel 2 turned off.")
    
    def run(self):
        self.window.mainloop()

def main():
    CPX = CPX400DP()
    CPX.autoConnect()
    GUI = GuiAppWithButtons(CPX)
    GUI.run()

if __name__ == '__main__':
    # Clear latest.log if it exists
    if os.path.exists('latest.log'):
        open('latest.log', 'w').close()
    
    # File handler
    file_handler = logging.FileHandler('latest.log', encoding='utf-8')
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(logging.Formatter('%(asctime)s.%(msecs)03d %(levelname)s: %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
    logger.addHandler(file_handler)
    
    # Console handler
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(logging.Formatter('%(asctime)s.%(msecs)03d %(levelname)s: %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
    logger.addHandler(console_handler)
    
    # Set the overall logging level
    logger.setLevel(logging.DEBUG)
    
    try:
        main()
    except Exception as e:
        logger.fatal(f'The script could no longer continue due to {repr(e)}.')
        input('Press any key to exit.')
        exit(1)
      
# interfata cu utilizatorul in care scrii de la tastatura ce valori vrei pt power supply

class GuiApp:
    def __init__(self, power_supply_instance, master=None):
        self.power_supply_instance = power_supply_instance
        
        # Build UI
        self.window = tk.Tk() if master is None else tk.Toplevel(master)
        self.window.configure(height=140, width=300)
        self.window.resizable(False, False)
        self.window.title("Power Supply Control")
        self.channel_1_label = ttk.Label(self.window)
        self.channel_1_label.configure(text='Channel 1')
        self.channel_1_label.place(anchor="w", x=50, y=20)
        self.channel_2_label = ttk.Label(self.window)
        self.channel_2_label.configure(text='Channel 2')
        self.channel_2_label.place(anchor="w", x=210, y=20)
        self.v_label_1 = ttk.Label(self.window)
        self.v_label_1.configure(text='V')
        self.v_label_1.place(anchor="center", x=35, y=40)
        self.i_label_1 = ttk.Label(self.window)
        self.i_label_1.configure(text='I')
        self.i_label_1.place(anchor="center", x=35, y=65)
        self.channel_1_voltage = ttk.Entry(self.window)
        self.ch1v = tk.DoubleVar()
        self.channel_1_voltage.configure(textvariable=self.ch1v, width=5)
        self.channel_1_voltage.place(anchor="w", x=50, y=40)
        self.channel_1_current = ttk.Entry(self.window)
        self.ch1i = tk.DoubleVar()
        self.channel_1_current.configure(textvariable=self.ch1i, width=5)
        self.channel_1_current.place(anchor="w", x=50, y=65)
        self.channel_2_voltage = ttk.Entry(self.window)
        self.ch2v = tk.DoubleVar()
        self.channel_2_voltage.configure(textvariable=self.ch2v, width=5)
        self.channel_2_voltage.place(anchor="w", x=210, y=40)
        self.channel_2_current = ttk.Entry(self.window)
        self.ch2i = tk.DoubleVar()
        self.channel_2_current.configure(textvariable=self.ch2i, width=5)
        self.channel_2_current.place(anchor="w", x=210, y=65)
        self.set_channel_1 = ttk.Button(self.window)
        self.set_channel_1.configure(text='Set', width=5)
        self.set_channel_1.place(anchor="w", x=48, y=90)
        self.set_channel_1.configure(command=self.update_channel_1)
        self.set_channel_2 = ttk.Button(self.window)
        self.set_channel_2.configure(text='Set', width=5)
        self.set_channel_2.place(anchor="w", x=208, y=90)
        self.set_channel_2.configure(command=self.update_channel_2)
        self.channel_1_enable = ttk.Checkbutton(self.window)
        self.enableChannel1 = tk.BooleanVar()
        self.channel_1_enable.configure(text='Enable', variable=self.enableChannel1)
        self.channel_1_enable.place(anchor="w", x=47, y=115)
        self.channel_1_enable.configure(command=self.toggle_channel_1)
        self.channel_2_enable = ttk.Checkbutton(self.window)
        self.enableChannel2 = tk.BooleanVar()
        self.channel_2_enable.configure(text='Enable', variable=self.enableChannel2)
        self.channel_2_enable.place(anchor="w", x=207, y=115)
        self.channel_2_enable.configure(command=self.toggle_channel_2)
        self.v_label_2 = ttk.Label(self.window)
        self.v_label_2.configure(text='V')
        self.v_label_2.place(anchor="center", x=195, y=40)
        self.i_label_2 = ttk.Label(self.window)
        self.i_label_2.configure(text='I')
        self.i_label_2.place(anchor="center", x=195, y=65)
        separator1 = ttk.Separator(self.window)
        separator1.configure(orient="vertical")
        separator1.place(anchor="center", height=100, relx=0.5, rely=0.5)
        
        self.mainwindow = self.window # Main widget
    
    def run(self):
        self.mainwindow.mainloop()
    
    def update_channel_1(self):
        voltage = self.ch1v.get()
        current = self.ch1i.get()
        self.power_supply_instance.setVoltage(voltage, channel = 1)
        self.power_supply_instance.setCurrent(current, channel = 1)
    
    def update_channel_2(self):
        voltage = self.ch2v.get()
        current = self.ch2i.get()
        self.power_supply_instance.setVoltage(voltage, channel = 2)
        self.power_supply_instance.setCurrent(current, channel = 2)
    
    def toggle_channel_1(self):
        state = self.enableChannel1.get()
        if state:
            self.power_supply_instance.enableOutput(channel = 1)
        else:
            self.power_supply_instance.disableOutput(channel = 1)
    
    def toggle_channel_2(self):
        state = self.enableChannel2.get()
        if state:
            self.power_supply_instance.enableOutput(channel = 2)
        else:
            self.power_supply_instance.disableOutput(channel = 2)

def main():
    CPX = CPX400DP()
    CPX.autoConnect()
    GUI = GuiApp(CPX)
    GUI.run()

if __name__ == '__main__':
    # Clear latest.log if it exists
    if os.path.exists('latest.log'):
        open('latest.log', 'w').close()
    
    # File handler
    file_handler = logging.FileHandler('latest.log', encoding='utf-8')
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(logging.Formatter('%(asctime)s.%(msecs)03d %(levelname)s: %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
    logger.addHandler(file_handler)
    
    # Console handler
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(logging.Formatter('%(asctime)s.%(msecs)03d %(levelname)s: %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
    logger.addHandler(console_handler)
    
    # Set the overall logging level
    logger.setLevel(logging.DEBUG)
    
    try:
        main()
    except Exception as e:
        logger.fatal(f'The script could no longer continue due to {repr(e)}.')
        input('Press any key to exit.')
        exit(1)